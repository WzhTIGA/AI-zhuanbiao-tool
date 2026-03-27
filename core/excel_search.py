from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from typing import Literal
import xml.etree.ElementTree as ET

import openpyxl


MatchMode = Literal["精准匹配", "包含匹配"]


@dataclass(frozen=True)
class NameHit:
    file_name: str
    sheet_name: str
    row: int
    id_value: str | None
    key_value: str | None
    b_value: str | None


@dataclass(frozen=True)
class SecondaryHit:
    file_name: str
    sheet_name: str
    row: int
    matched_by: str
    source_id: str | None
    source_key: str | None
    source_b: str | None
    row_data: dict[str, str]


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_header(v: object) -> str:
    return _cell_str(v).casefold()


_TYPE_TOKENS = {
    "int",
    "integer",
    "bigint",
    "smallint",
    "tinyint",
    "float",
    "double",
    "real",
    "decimal",
    "number",
    "string",
    "str",
    "text",
    "varchar",
    "char",
    "numeric",
    "bool",
    "boolean",
    "date",
    "datetime",
    "timestamp",
}


def _is_type_desc(s: str) -> bool:
    t = s.casefold().strip()
    if not t:
        return False
    if t in _TYPE_TOKENS:
        return True
    if t.startswith("varchar(") or t.startswith("decimal("):
        return True
    if t.startswith("uint") and t[4:].isdigit():
        return True
    if t.startswith("int") and t[3:].isdigit():
        return True
    return False


def _row_span(ws, row: int, *, max_col: int) -> tuple[int, int] | None:
    max_scan = min(max(max_col, 1), 2048)
    first: int | None = None
    last: int | None = None
    empty_run = 0
    row_values = next(
        ws.iter_rows(min_row=row, max_row=row, min_col=1, max_col=max_scan, values_only=True),
        (),
    )
    for col, raw in enumerate(row_values, start=1):
        if _cell_str(raw):
            if first is None:
                first = col
            last = col
            empty_run = 0
        else:
            if first is not None:
                empty_run += 1
                if empty_run >= 50:
                    break
    if first is None or last is None:
        return None
    return first, last


def _header_span(ws, *, header_row: int = 1) -> tuple[int, int] | None:
    return _row_span(ws, header_row, max_col=ws.max_column or 1)


def _is_header_candidate(ws, row: int, *, max_col: int) -> bool:
    span = _row_span(ws, row, max_col=max_col)
    if span is None:
        return False
    first, last = span
    values = list(
        next(
            ws.iter_rows(min_row=row, max_row=row, min_col=first, max_col=last, values_only=True),
            (),
        )
    )
    values = [_cell_str(v) for v in values]
    if any(not v for v in values):
        return False
    norm = [v.casefold() for v in values]
    if len(set(norm)) != len(norm):
        return False
    if values and sum(1 for v in values if _is_type_desc(v)) / len(values) >= 0.6:
        return False
    return True


def _find_header_row(ws, *, max_scan_rows: int = 20) -> int:
    scan_to = min(max_scan_rows, ws.max_row or 1)
    max_col = max(ws.max_column or 1, 1)
    best_row: int | None = None
    best_width = 0
    for row in range(1, scan_to + 1):
        if not _is_header_candidate(ws, row, max_col=max_col):
            continue
        span = _row_span(ws, row, max_col=max_col)
        if span is None:
            continue
        first, last = span
        width = last - first + 1
        if width > best_width:
            best_row = row
            best_width = width
    return best_row or 1


def _is_processed_sheet(ws) -> bool:
    return (
        _cell_str(ws.cell(row=1, column=1).value) == "##var"
        and _cell_str(ws.cell(row=2, column=1).value) == "##type"
        and _cell_str(ws.cell(row=3, column=1).value) == "##"
    )


def _is_type_row(ws, header_span: tuple[int, int], *, header_row: int) -> bool:
    row = header_row + 1
    if row > (ws.max_row or 0):
        return False
    first, last = header_span
    non_empty = 0
    hits = 0
    row_values = next(
        ws.iter_rows(min_row=row, max_row=row, min_col=first, max_col=last, values_only=True),
        (),
    )
    for raw in row_values:
        v = _cell_str(raw)
        if not v:
            continue
        non_empty += 1
        if _is_type_desc(v):
            hits += 1
    if non_empty == 0:
        return False
    return hits / non_empty >= 0.6


@dataclass(frozen=True)
class _SheetLayout:
    processed: bool
    header_row: int
    header_span: tuple[int, int]
    data_start_row: int


def _sheet_layout(ws) -> _SheetLayout:
    processed = _is_processed_sheet(ws)
    if processed:
        header_row = 1
    else:
        header_row = _find_header_row(ws, max_scan_rows=20)

    max_col = max(ws.max_column or 1, 1)
    span = _row_span(ws, header_row, max_col=max_col) or (1, min(max_col, 2048))

    if processed:
        data_start_row = 4
    else:
        data_start_row = header_row + 1 + (1 if _is_type_row(ws, span, header_row=header_row) else 0)
    return _SheetLayout(processed=processed, header_row=header_row, header_span=span, data_start_row=data_start_row)


def _row_has_any_in_cols(ws, *, row: int, cols: list[int]) -> bool:
    if not cols:
        return False
    min_col = min(cols)
    max_col = max(cols)
    row_values = next(
        ws.iter_rows(min_row=row, max_row=row, min_col=min_col, max_col=max_col, values_only=True),
        (),
    )
    for col in cols:
        raw = row_values[col - min_col] if (min_col <= col <= max_col) else None
        if _cell_str(raw):
            return True
    return False


def _find_data_window(ws, *, start_row: int, cols: list[int]) -> tuple[int, int] | None:
    if not cols:
        return None

    max_row = ws.max_row or 0
    if start_row > max_row:
        return None

    r = start_row
    for _ in range(5000):
        if r > max_row:
            return None
        if _row_has_any_in_cols(ws, row=r, cols=cols):
            break
        r += 1
    else:
        return None

    start = r
    empty_run = 0
    scanned = 0
    while r <= max_row and scanned < 200000:
        if _row_has_any_in_cols(ws, row=r, cols=cols):
            empty_run = 0
        else:
            empty_run += 1
            if empty_run >= 200:
                end = r - empty_run
                return (start, end) if end >= start else None
        r += 1
        scanned += 1
    return (start, r - 1) if (r - 1) >= start else None


def _find_col_by_header(ws, header_name: str, header_row: int = 1) -> int | None:
    target = header_name.casefold().strip()
    span = _row_span(ws, header_row, max_col=ws.max_column or 1)
    if span is None:
        first = 1
        last = min(ws.max_column or 1, 2048)
    else:
        first, last = span
    row_values = next(
        ws.iter_rows(min_row=header_row, max_row=header_row, min_col=first, max_col=last, values_only=True),
        (),
    )
    for offset, raw in enumerate(row_values):
        if _norm_header(raw) == target:
            return first + offset
    return None


def count_sheets_in_xlsx_bytes(data: bytes) -> int:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml_bytes = zf.read("xl/workbook.xml")
        root = ET.fromstring(xml_bytes)
        return sum(1 for el in root.iter() if el.tag.rsplit("}", 1)[-1] == "sheet")
    except Exception:
        return 0


def _read_row_values(ws, *, row: int, min_col: int, max_col: int) -> tuple[object, ...]:
    return next(
        ws.iter_rows(min_row=row, max_row=row, min_col=min_col, max_col=max_col, values_only=True),
        (),
    )


def _scan_column_for_match(
    ws,
    *,
    col: int,
    start_row: int,
    match_mode: MatchMode,
    query: str,
    max_leading_skip: int = 5000,
    max_scan_rows: int = 200000,
    empty_run_limit: int = 200,
) -> int | None:
    q = str(query).strip()
    if not q:
        return None
    max_row = ws.max_row or 0
    if start_row > max_row:
        return None

    started = False
    leading = 0
    empty_run = 0
    scanned = 0
    row = start_row

    it = ws.iter_rows(min_row=start_row, max_row=max_row, min_col=col, max_col=col, values_only=True)
    for (raw,) in it:
        scanned += 1
        v = _cell_str(raw)

        if not started:
            if v:
                started = True
            else:
                leading += 1
                if leading >= max_leading_skip:
                    return None
                row += 1
                if scanned >= max_scan_rows:
                    return None
                continue

        if not v:
            empty_run += 1
            if empty_run >= empty_run_limit:
                return None
        else:
            empty_run = 0
            ok = v == q if match_mode == "精准匹配" else (q in v)
            if ok:
                return row

        row += 1
        if scanned >= max_scan_rows:
            return None
    return None


def _scan_column_for_exact(
    ws,
    *,
    col: int,
    start_row: int,
    query: str,
    max_leading_skip: int = 5000,
    max_scan_rows: int = 200000,
    empty_run_limit: int = 200,
) -> int | None:
    q = str(query).strip()
    if not q:
        return None
    max_row = ws.max_row or 0
    if start_row > max_row:
        return None

    started = False
    leading = 0
    empty_run = 0
    scanned = 0
    row = start_row

    it = ws.iter_rows(min_row=start_row, max_row=max_row, min_col=col, max_col=col, values_only=True)
    for (raw,) in it:
        scanned += 1
        v = _cell_str(raw)

        if not started:
            if v:
                started = True
            else:
                leading += 1
                if leading >= max_leading_skip:
                    return None
                row += 1
                if scanned >= max_scan_rows:
                    return None
                continue

        if not v:
            empty_run += 1
            if empty_run >= empty_run_limit:
                return None
        else:
            empty_run = 0
            if v == q:
                return row

        row += 1
        if scanned >= max_scan_rows:
            return None
    return None


def _build_row_data(ws, row: int, layout: _SheetLayout) -> dict[str, str]:
    out: dict[str, str] = {}
    first, last = layout.header_span
    start_col = 2 if layout.processed else first
    for col in range(start_col, last + 1):
        if layout.processed:
            key = _cell_str(ws.cell(row=3, column=col).value) or _cell_str(ws.cell(row=1, column=col).value)
        else:
            key = _cell_str(ws.cell(row=layout.header_row, column=col).value)
        if not key:
            continue
        out[key] = _cell_str(ws.cell(row=row, column=col).value)
    return out


def find_name_hits(
    *,
    workbooks: dict[str, bytes],
    query_name: str,
    match_mode: MatchMode,
    progress_cb=None,
    sheet_total: int | None = None,
) -> list[NameHit]:
    q = str(query_name).strip()
    if not q:
        return []

    hits: list[NameHit] = []
    current_sheet = 0
    total = int(sheet_total or 0)
    for file_name, data in workbooks.items():
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                current_sheet += 1
                if progress_cb is not None:
                    progress_cb(current_sheet, total, file_name, ws.title)
                layout = _sheet_layout(ws)
                name_col = _find_col_by_header(ws, "name", header_row=layout.header_row)
                id_col = _find_col_by_header(ws, "id", header_row=layout.header_row)
                key_col = _find_col_by_header(ws, "key", header_row=layout.header_row)
                if name_col is not None:
                    row = _scan_column_for_match(
                        ws,
                        col=name_col,
                        start_row=layout.data_start_row,
                        match_mode=match_mode,
                        query=q,
                    )
                    if row is not None:
                        cols = [c for c in [id_col, key_col, 2] if isinstance(c, int)]
                        if cols:
                            min_c = min(cols)
                            max_c = max(cols)
                            row_vals = _read_row_values(ws, row=row, min_col=min_c, max_col=max_c)
                            id_value = (
                                _cell_str(row_vals[id_col - min_c]) if id_col and (min_c <= id_col <= max_c) else ""
                            )
                            key_value = (
                                _cell_str(row_vals[key_col - min_c])
                                if key_col and (min_c <= key_col <= max_c)
                                else ""
                            )
                            b_value = _cell_str(row_vals[2 - min_c]) if (min_c <= 2 <= max_c) else ""
                        else:
                            id_value = ""
                            key_value = ""
                            b_value = ""

                        hits.append(
                            NameHit(
                                file_name=file_name,
                                sheet_name=ws.title,
                                row=row,
                                id_value=id_value or None,
                                key_value=key_value or None,
                                b_value=b_value or None,
                            )
                        )
                else:
                    found = False
                    _, max_col = layout.header_span
                    min_col = 2 if layout.processed else layout.header_span[0]
                    empty_run = 0
                    leading = 0
                    started = False
                    scanned = 0
                    max_row = ws.max_row or 0
                    for row, row_values in enumerate(
                        ws.iter_rows(
                            min_row=layout.data_start_row,
                            max_row=max_row,
                            min_col=min_col,
                            max_col=max_col,
                            values_only=True,
                        ),
                        start=layout.data_start_row,
                    ):
                        scanned += 1
                        if scanned >= 200000:
                            break

                        has_any = False
                        for raw in row_values:
                            v = _cell_str(raw)
                            if not v:
                                continue
                            has_any = True
                            started = True
                            empty_run = 0
                            ok = v == q if match_mode == "精准匹配" else (q in v)
                            if ok:
                                cols = [c for c in [id_col, key_col, 2] if isinstance(c, int)]
                                if cols:
                                    min_c = min(cols)
                                    max_c = max(cols)
                                    row_vals = _read_row_values(ws, row=row, min_col=min_c, max_col=max_c)
                                    id_value = (
                                        _cell_str(row_vals[id_col - min_c])
                                        if id_col and (min_c <= id_col <= max_c)
                                        else ""
                                    )
                                    key_value = (
                                        _cell_str(row_vals[key_col - min_c])
                                        if key_col and (min_c <= key_col <= max_c)
                                        else ""
                                    )
                                    b_value = _cell_str(row_vals[2 - min_c]) if (min_c <= 2 <= max_c) else ""
                                else:
                                    id_value = ""
                                    key_value = ""
                                    b_value = ""

                                hits.append(
                                    NameHit(
                                        file_name=file_name,
                                        sheet_name=ws.title,
                                        row=row,
                                        id_value=id_value or None,
                                        key_value=key_value or None,
                                        b_value=b_value or None,
                                    )
                                )
                                found = True
                                break
                        if found:
                            break
                        if not started:
                            if not has_any:
                                leading += 1
                                if leading >= 5000:
                                    break
                            continue
                        if not has_any:
                            empty_run += 1
                            if empty_run >= 200:
                                break
        finally:
            wb.close()

    return hits


def find_secondary_hits(
    *,
    workbooks: dict[str, bytes],
    source: NameHit,
    only_files: set[str] | None = None,
    progress_cb=None,
    sheet_total: int | None = None,
) -> list[SecondaryHit]:
    src_id = source.id_value
    src_key = source.key_value
    src_b = source.b_value

    hits: list[SecondaryHit] = []
    current_sheet = 0
    total = int(sheet_total or 0)
    for file_name, data in workbooks.items():
        if only_files is not None and file_name not in only_files:
            continue
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                current_sheet += 1
                if progress_cb is not None:
                    progress_cb(current_sheet, total, file_name, ws.title)
                layout = _sheet_layout(ws)
                id_col = _find_col_by_header(ws, "id", header_row=layout.header_row)
                key_col = _find_col_by_header(ws, "key", header_row=layout.header_row)
                row: int | None = None
                matched_by: str | None = None

                if src_id and id_col is not None:
                    row = _scan_column_for_exact(ws, col=id_col, start_row=layout.data_start_row, query=src_id)
                    matched_by = "id" if row is not None else None

                if row is None and src_key and key_col is not None:
                    row = _scan_column_for_exact(ws, col=key_col, start_row=layout.data_start_row, query=src_key)
                    matched_by = "key" if row is not None else None

                if row is None and src_b:
                    row = _scan_column_for_exact(ws, col=2, start_row=layout.data_start_row, query=src_b)
                    matched_by = "B列" if row is not None else None

                if row is not None and matched_by is not None:
                    hits.append(
                        SecondaryHit(
                            file_name=file_name,
                            sheet_name=ws.title,
                            row=row,
                            matched_by=matched_by,
                            source_id=src_id,
                            source_key=src_key,
                            source_b=src_b,
                            row_data=_build_row_data(ws, row, layout),
                        )
                    )
        finally:
            wb.close()

    return hits
