from __future__ import annotations

import io
import datetime as _dt
import re
from dataclasses import dataclass

import openpyxl

from .translate.base import Translator


_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1F]')
_INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")


def sanitize_windows_filename(name: str) -> str:
    name = _INVALID_FILENAME_CHARS.sub("_", name).strip()
    name = name.rstrip(". ")
    if not name:
        return "未命名"
    return name


def filename_base(name: str) -> str:
    lower = name.lower()
    if lower.endswith(".xlsx"):
        return name[: -len(".xlsx")]
    return name


def sanitize_sheet_title(title: str) -> str:
    title = _INVALID_SHEET_CHARS.sub(" ", str(title)).strip()
    title = " ".join(title.split())
    if not title:
        return "Sheet"
    if len(title) > 31:
        return title[:31]
    return title


def make_unique_sheet_title(*, desired: str, used: set[str]) -> str:
    base = sanitize_sheet_title(desired)
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 2000):
        suffix = f"({i})"
        max_len = 31 - len(suffix)
        cand = (base[:max_len] if max_len > 0 else "") + suffix
        if cand not in used:
            used.add(cand)
            return cand
    fallback = "Sheet"
    used.add(fallback)
    return fallback


def _cell_text(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _row_span(ws, row: int, *, max_col: int) -> tuple[int, int] | None:
    max_scan = min(max_col, 2048)
    first: int | None = None
    last: int | None = None
    empty_run = 0
    for col in range(1, max_scan + 1):
        if _cell_text(ws.cell(row=row, column=col).value):
            if first is None:
                first = col
            last = col
            empty_run = 0
        else:
            if first is not None:
                empty_run += 1
                if empty_run >= 50:
                    break
    if first is None or last is None:
        return None
    return first, last


_TYPE_TOKENS: set[str] = {
    "int",
    "integer",
    "bigint",
    "smallint",
    "str",
    "string",
    "text",
    "char",
    "varchar",
    "float",
    "double",
    "decimal",
    "number",
    "numeric",
    "bool",
    "boolean",
    "date",
    "datetime",
    "timestamp",
}


def _is_type_desc(s: str) -> bool:
    t = s.casefold().strip()
    if not t:
        return False
    if t in _TYPE_TOKENS:
        return True
    if t.startswith("varchar(") or t.startswith("decimal("):
        return True
    if t.startswith("uint") and t[4:].isdigit():
        return True
    if t.startswith("int") and t[3:].isdigit():
        return True
    return False


def _is_header_candidate(ws, row: int, *, max_col: int) -> bool:
    span = _row_span(ws, row, max_col=max_col)
    if span is None:
        return False
    first, last = span
    values = [_cell_text(ws.cell(row=row, column=col).value) for col in range(first, last + 1)]
    if any(not v for v in values):
        return False
    norm = [v.casefold() for v in values]
    if len(set(norm)) != len(norm):
        return False
    if values and sum(1 for v in values if _is_type_desc(v)) / len(values) >= 0.6:
        return False
    return True


def _find_header_row(ws, *, max_scan_rows: int = 20) -> int:
    scan_to = min(max_scan_rows, ws.max_row or 1)
    max_col = max(ws.max_column or 1, 1)
    best_row: int | None = None
    best_width = 0
    for row in range(1, scan_to + 1):
        if not _is_header_candidate(ws, row, max_col=max_col):
            continue
        span = _row_span(ws, row, max_col=max_col)
        if span is None:
            continue
        first, last = span
        width = last - first + 1
        if width > best_width:
            best_row = row
            best_width = width
    return best_row or 1


def _is_type_row(ws, header_span: tuple[int, int], *, header_row: int = 1) -> bool:
    row = header_row + 1
    if row > (ws.max_row or 0):
        return False
    first, last = header_span
    non_empty = 0
    hits = 0
    for col in range(first, last + 1):
        v = _cell_text(ws.cell(row=row, column=col).value)
        if not v:
            continue
        non_empty += 1
        if _is_type_desc(v):
            hits += 1
    if non_empty == 0:
        return False
    return hits / non_empty >= 0.6


def _row_has_any(ws, row: int, span: tuple[int, int]) -> bool:
    first, last = span
    for col in range(first, last + 1):
        if _cell_text(ws.cell(row=row, column=col).value):
            return True
    return False


def _find_data_block(ws, *, start_row: int, span: tuple[int, int]) -> tuple[int, int] | None:
    max_row = ws.max_row or 0
    r = start_row
    while r <= max_row and not _row_has_any(ws, r, span):
        r += 1
    if r > max_row:
        return None
    start = r
    end = r
    r += 1
    while r <= max_row and _row_has_any(ws, r, span):
        end = r
        r += 1
    return start, end


def _infer_cell_kind(v: object) -> str | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, _dt.datetime):
        return "datetime"
    if isinstance(v, _dt.date):
        return "date"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float" if v.is_integer() else "float"  
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        sl = s.casefold()
        if sl in {"true", "false"}:
            return "bool"
        if re.fullmatch(r"[-+]?\d+", s):
            return "int"
        if re.fullmatch(r"[-+]?\d+\.\d+", s):
            return "float"
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", s):
            return "date"
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}[ T]\d{1,2}:\d{2}(:\d{2})?", s):
            return "datetime"
        return "string"
    return "string"


def _infer_col_type(ws, *, col: int, data_block: tuple[int, int] | None) -> str:
    if data_block is None:
        return "string"
    start, end = data_block
    kinds: list[str] = []
    for row in range(start, min(end, start + 49) + 1):
        k = _infer_cell_kind(ws.cell(row=row, column=col).value)
        if k is None:
            continue
        kinds.append(k)
    if not kinds:
        return "string"
    if "string" in kinds:
        return "string"
    if "datetime" in kinds:
        return "datetime"
    if "date" in kinds:
        return "date"
    if "float" in kinds:
        return "float"
    if set(kinds) == {"bool"}:
        return "bool"
    if set(kinds) == {"int"}:
        return "int"
    return "string"


@dataclass(frozen=True)
class ProcessedWorkbook:
    original_filename: str
    original_base: str
    translated_base: str
    output_filename: str
    output_bytes: bytes


def process_workbook(
    *,
    original_filename: str,
    xlsx_bytes: bytes,
    translator: Translator,
    src_lang: str = "en",
    dst_lang: str = "zh-CN",
) -> ProcessedWorkbook:
    original_base = filename_base(original_filename)
    translated_base = translator.translate(text=original_base, src_lang=src_lang, dst_lang=dst_lang)
    safe_translated = sanitize_windows_filename(translated_base)
    safe_original = sanitize_windows_filename(original_base)
    output_filename = f"{safe_original}({safe_translated}).xlsx"

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    try:
        used_titles: set[str] = set()
        for ws in wb.worksheets:
            desired_title = translator.translate(text=ws.title, src_lang=src_lang, dst_lang=dst_lang)
            ws.title = make_unique_sheet_title(desired=desired_title, used=used_titles)

            already = (
                _cell_text(ws.cell(row=1, column=1).value) == "##var"
                and _cell_text(ws.cell(row=2, column=1).value) == "##type"
                and _cell_text(ws.cell(row=3, column=1).value) == "##"
            )

            if not already:
                header_row = _find_header_row(ws, max_scan_rows=20)
                if header_row > 1:
                    ws.delete_rows(1, header_row - 1)

                max_col = max(ws.max_column or 1, 1)
                span0 = _row_span(ws, 1, max_col=max_col) or (1, max_col)
                has_type = _is_type_row(ws, span0, header_row=1)
                inserted_type = False
                if not has_type:
                    ws.insert_rows(2)
                    inserted_type = True
                ws.insert_rows(3)
                ws.insert_cols(1)

                ws.cell(row=1, column=1).value = "##var"
                ws.cell(row=2, column=1).value = "##type"
                ws.cell(row=3, column=1).value = "##"

                span = (span0[0] + 1, span0[1] + 1)
                max_row = ws.max_row or 0
                if max_row >= 4:
                    r = 4
                    while r <= max_row and not _row_has_any(ws, r, span):
                        r += 1
                    empty_count = r - 4
                    if empty_count > 0:
                        ws.delete_rows(4, empty_count)

                data_block = _find_data_block(ws, start_row=4, span=span)
                if inserted_type:
                    for col in range(span[0], span[1] + 1):
                        t = _infer_col_type(ws, col=col, data_block=data_block)
                        ws.cell(row=2, column=col).value = t
                else:
                    for col in range(span[0], span[1] + 1):
                        if _cell_text(ws.cell(row=2, column=col).value):
                            continue
                        t = _infer_col_type(ws, col=col, data_block=data_block)
                        ws.cell(row=2, column=col).value = t
            else:
                max_col = max(ws.max_column or 1, 1)
                span = _row_span(ws, 1, max_col=max_col) or (2, max_col)
                if span[0] == 1:
                    span = (2, span[1])
                max_row = ws.max_row or 0
                if max_row >= 4:
                    r = 4
                    while r <= max_row and not _row_has_any(ws, r, span):
                        r += 1
                    empty_count = r - 4
                    if empty_count > 0:
                        ws.delete_rows(4, empty_count)
                data_block = _find_data_block(ws, start_row=4, span=span)
                for col in range(span[0], span[1] + 1):
                    if _cell_text(ws.cell(row=2, column=col).value):
                        continue
                    t = _infer_col_type(ws, col=col, data_block=data_block)
                    ws.cell(row=2, column=col).value = t

            local: dict[str, str] = {}
            for col in range(span[0], span[1] + 1):
                v = ws.cell(row=1, column=col).value
                if v is None:
                    continue
                header = str(v).strip()
                if not header:
                    continue
                zh = local.get(header)
                if zh is None:
                    zh = translator.translate(text=header, src_lang=src_lang, dst_lang=dst_lang)
                    local[header] = zh
                ws.cell(row=3, column=col).value = zh

        out_buf = io.BytesIO()
        wb.save(out_buf)
        output_bytes = out_buf.getvalue()
    finally:
        wb.close()

    return ProcessedWorkbook(
        original_filename=original_filename,
        original_base=original_base,
        translated_base=translated_base,
        output_filename=output_filename,
        output_bytes=output_bytes,
    )
