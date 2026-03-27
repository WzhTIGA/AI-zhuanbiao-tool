from __future__ import annotations

import io
import zipfile

import openpyxl

from core.pipeline import default_type_dict, run_pipeline
from core.translate.base import Translator


class FakeTranslator(Translator):
    def translate(self, *, text: str, src_lang: str, dst_lang: str) -> str:
        t = str(text).strip()
        if not t:
            return t
        return f"中_{t}"


def _make_sample_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主表"
    ws["A1"] = "id"
    ws["B1"] = "key"
    ws["C1"] = "name"
    ws["A4"] = 1
    ws["B4"] = "k1"
    ws["C4"] = "测试道具"

    ws2 = wb.create_sheet("附表1")
    ws2["A1"] = "说明行"
    ws2["A3"] = "id"
    ws2["B3"] = "flag"
    ws2["C3"] = "created_at"
    ws2["A5"] = 2
    ws2["B5"] = "true"
    ws2["C5"] = "2026-03-24"

    ws3 = wb.create_sheet("附表2")
    ws3["A1"] = "id"
    ws3["B1"] = "price"
    ws3["A2"] = "int"
    ws3["B2"] = "float"
    ws3["A4"] = 3
    ws3["B4"] = 12.5

    ws4 = wb.create_sheet("膨胀表")
    ws4["A1"] = "id"
    ws4["B1"] = "name"
    ws4["A4"] = 9
    ws4["B4"] = "无关数据"
    ws4.cell(row=200000, column=3000).value = ""

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _make_zip() -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("item_testTable.xlsx", _make_sample_xlsx())
    return buf.getvalue()


def main() -> None:
    out = run_pipeline(
        zip_bytes=_make_zip(),
        query_type_en="item",
        query_name="测试道具",
        match_mode="精准匹配",
        type_dict=default_type_dict(),
        translator=FakeTranslator(),
    )

    assert out.processed_zip_bytes
    assert out.summary["处理文件数"] == 2
    assert out.summary["第4步_名称搜索"]["状态"] == "已命中"
    assert out.summary["第6步_同类二次检索"]["状态"] == "已命中"

    zf = zipfile.ZipFile(io.BytesIO(out.processed_zip_bytes))
    names = zf.namelist()
    assert len(names) == 2
    assert "命中数据.xlsx" in names
    processed_name = [n for n in names if n != "命中数据.xlsx"][0]
    data = zf.read(processed_name)
    zf.close()

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.worksheets[0]
    assert ws["A1"].value == "##var"
    assert ws["A2"].value == "##type"
    assert ws["A3"].value == "##"
    assert ws["B3"].value == "中_id"
    assert ws["B2"].value == "int"
    assert ws["C2"].value == "str"
    assert ws["D2"].value == "str"
    assert ws["D4"].value == "测试道具"

    ws2 = wb.worksheets[1]
    assert ws2["A1"].value == "##var"
    assert ws2["A2"].value == "##type"
    assert ws2["A3"].value == "##"
    assert ws2["B1"].value == "id"
    assert ws2["C1"].value == "flag"
    assert ws2["D1"].value == "created_at"
    assert ws2["B2"].value == "int"
    assert ws2["C2"].value == "bool"
    assert ws2["D2"].value == "date"
    assert ws2["B3"].value == "中_id"
    assert ws2["D4"].value == "2026-03-24"

    ws3 = wb.worksheets[2]
    assert ws3["A1"].value == "##var"
    assert ws3["A2"].value == "##type"
    assert ws3["A3"].value == "##"
    assert ws3["B1"].value == "id"
    assert ws3["C1"].value == "price"
    assert ws3["B2"].value == "int"
    assert ws3["C2"].value == "float"
    assert ws3["B3"].value == "中_id"
    assert ws3["C4"].value == 12.5

    ws4 = wb.worksheets[3]
    assert ws4["A1"].value == "##var"
    assert ws4["A2"].value == "##type"
    assert ws4["A3"].value == "##"
    assert ws4["B1"].value == "id"
    assert ws4["C1"].value == "name"
    assert ws4["B2"].value == "int"
    assert ws4["C2"].value == "str"
    wb.close()


if __name__ == "__main__":
    main()
